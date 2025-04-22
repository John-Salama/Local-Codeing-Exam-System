import os
import time
import sqlite3
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import secrets
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

# Add chr function to Jinja2 environment
app.jinja_env.globals.update(chr=chr)

# Function to get real IP address (handles X-Forwarded-For for testing)


def get_real_ip():
    """
    Gets the real IP address of the client, taking into account X-Forwarded-For header
    which is used by the stress testing script for IP rotation.
    """
    if 'X-Forwarded-For' in request.headers:
        return request.headers.get('X-Forwarded-For')
    return request.remote_addr


# Database setup
DATABASE_PATH = 'database/exam_system.db'


def init_db():
    """Initialize database with required tables"""
    with sqlite3.connect(DATABASE_PATH) as conn:
        cursor = conn.cursor()

        # Create users table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        )
        ''')

        # Create exams table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS exams (
            id INTEGER PRIMARY KEY,
            title TEXT NOT NULL,
            duration INTEGER NOT NULL,
            is_active BOOLEAN DEFAULT 0,
            created_by INTEGER,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
        ''')

        # Create exam_models table for multiple exam variants
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS exam_models (
            id INTEGER PRIMARY KEY,
            exam_id INTEGER NOT NULL,
            model_name TEXT NOT NULL,
            FOREIGN KEY (exam_id) REFERENCES exams (id)
        )
        ''')

        # Create questions table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY,
            exam_id INTEGER,
            model_id INTEGER,
            question_text TEXT NOT NULL,
            FOREIGN KEY (exam_id) REFERENCES exams (id),
            FOREIGN KEY (model_id) REFERENCES exam_models (id)
        )
        ''')

        # Create submissions table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS submissions (
            id INTEGER PRIMARY KEY,
            student_id INTEGER,
            student_name TEXT NOT NULL,
            student_number TEXT NOT NULL,
            exam_id INTEGER,
            model_id INTEGER,
            submission_time TIMESTAMP,
            ip_address TEXT,
            code_content TEXT,
            version INTEGER,
            FOREIGN KEY (student_id) REFERENCES users (id),
            FOREIGN KEY (exam_id) REFERENCES exams (id),
            FOREIGN KEY (model_id) REFERENCES exam_models (id)
        )
        ''')

        # Create question_answers table for individual answers per question
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS question_answers (
            id INTEGER PRIMARY KEY,
            submission_id INTEGER,
            question_id INTEGER,
            answer_content TEXT,
            FOREIGN KEY (submission_id) REFERENCES submissions (id),
            FOREIGN KEY (question_id) REFERENCES questions (id)
        )
        ''')

        # Create ip_restrictions table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS ip_restrictions (
            id INTEGER PRIMARY KEY,
            ip_address TEXT NOT NULL UNIQUE,
            is_blocked BOOLEAN DEFAULT 0,
            blocked_time TIMESTAMP,
            approved BOOLEAN DEFAULT 0
        )
        ''')

        # Create exam_sessions table for tracking active exams
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS exam_sessions (
            id INTEGER PRIMARY KEY,
            student_name TEXT NOT NULL,
            student_number TEXT NOT NULL,
            exam_id INTEGER,
            model_id INTEGER,
            start_time TIMESTAMP,
            end_time TIMESTAMP,
            ip_address TEXT,
            FOREIGN KEY (exam_id) REFERENCES exams (id),
            FOREIGN KEY (model_id) REFERENCES exam_models (id)
        )
        ''')

        # Create grades table
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS grades (
            id INTEGER PRIMARY KEY,
            submission_id INTEGER,
            mark FLOAT,
            comment TEXT,
            graded_by INTEGER,
            graded_at TIMESTAMP,
            FOREIGN KEY (submission_id) REFERENCES submissions (id),
            FOREIGN KEY (graded_by) REFERENCES users (id)
        )
        ''')

        # Insert default teacher account if not exists
        cursor.execute("SELECT * FROM users WHERE username = 'teacher'")
        if not cursor.fetchone():
            teacher_password = generate_password_hash('admin123')
            cursor.execute(
                "INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                ('teacher', teacher_password, 'teacher')
            )

        conn.commit()


# Ensure database directory exists
os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)
init_db()

# Login route


@app.route('/')
def login():
    # Redirect root URL to the student login page
    return redirect(url_for('student_login'))


@app.route('/teacher_login', methods=['GET', 'POST'])
def teacher_login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM users WHERE username = ?", (username,))
            user = cursor.fetchone()

            if user and check_password_hash(user[2], password):
                session['user_id'] = user[0]
                session['username'] = user[1]
                session['role'] = user[3]

                if user[3] == 'teacher':
                    return redirect(url_for('teacher_dashboard'))
                else:
                    return redirect(url_for('student_login'))
            else:
                error = 'Invalid credentials'

    return render_template('login.html', error=error)

# Student login route (separate from teacher login)


@app.route('/student_login', methods=['GET', 'POST'])
def student_login():
    error = None
    if request.method == 'POST':
        name = request.form['name']
        student_number = request.form['student_number']
        exam_id = request.form.get('exam_id')

        if not name or not student_number:
            error = 'Please enter both name and student number'
            return render_template('student_login.html', error=error)

        # Check if an exam is active
        with sqlite3.connect(DATABASE_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # Get active exam
            cursor.execute("SELECT * FROM exams WHERE is_active = 1")
            exam = cursor.fetchone()

            if not exam:
                error = 'No active exam available'
                return render_template('student_login.html', error=error)

            # Check IP address
            ip_address = get_real_ip()
            cursor.execute(
                "SELECT * FROM ip_restrictions WHERE ip_address = ?", (ip_address,))
            ip_restriction = cursor.fetchone()

            if ip_restriction and ip_restriction['is_blocked'] and not ip_restriction['approved']:
                error = 'Your IP address is blocked. Please contact the teacher.'
                return render_template('student_login.html', error=error)

            # Check for existing session
            cursor.execute(
                """
                SELECT * FROM exam_sessions 
                WHERE student_number = ? AND exam_id = ? AND end_time > ?
                """,
                (student_number, exam['id'], datetime.now())
            )
            existing_session = cursor.fetchone()

            if not existing_session:
                # Get available models for this exam
                cursor.execute(
                    "SELECT * FROM exam_models WHERE exam_id = ?",
                    (exam['id'],)
                )
                available_models = cursor.fetchall()

                if not available_models:
                    # Create a default model if none exists (for backwards compatibility)
                    cursor.execute(
                        "INSERT INTO exam_models (exam_id, model_name) VALUES (?, ?)",
                        (exam['id'], "Default Model")
                    )
                    model_id = cursor.lastrowid
                else:
                    # Randomly select a model from available models
                    import random
                    model = random.choice(available_models)
                    model_id = model['id']

                # Create new exam session
                start_time = datetime.now()
                end_time = start_time + timedelta(minutes=exam['duration'])

                cursor.execute(
                    """
                    INSERT INTO exam_sessions 
                    (student_name, student_number, exam_id, model_id, start_time, end_time, ip_address) 
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (name, student_number, exam['id'], model_id,
                     start_time, end_time, ip_address)
                )
                session_id = cursor.lastrowid
                session['model_id'] = model_id
            else:
                session_id = existing_session['id']
                session['model_id'] = existing_session['model_id']

            session['student_name'] = name
            session['student_number'] = student_number
            session['exam_id'] = exam['id']
            session['session_id'] = session_id

            # Record IP address
            if not ip_restriction:
                cursor.execute(
                    "INSERT INTO ip_restrictions (ip_address, is_blocked, approved) VALUES (?, 0, 1)",
                    (ip_address,)
                )

            return redirect(url_for('take_exam'))

    return render_template('student_login.html', error=error)

# Student exam page


@app.route('/take_exam')
def take_exam():
    if 'student_name' not in session or 'student_number' not in session or 'exam_id' not in session or 'model_id' not in session:
        return redirect(url_for('student_login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get exam details
        cursor.execute("SELECT * FROM exams WHERE id = ?",
                       (session['exam_id'],))
        exam = cursor.fetchone()

        # Get exam model
        cursor.execute("SELECT * FROM exam_models WHERE id = ?",
                       (session['model_id'],))
        model = cursor.fetchone()

        # Get questions for this specific model
        cursor.execute("""
            SELECT * FROM questions 
            WHERE exam_id = ? AND model_id = ?
            ORDER BY id
            """,
                       (session['exam_id'], session['model_id'])
                       )
        questions = cursor.fetchall()

        # Get student's session info
        cursor.execute("SELECT * FROM exam_sessions WHERE id = ?",
                       (session['session_id'],))
        exam_session = cursor.fetchone()

        if not exam_session:
            return redirect(url_for('student_login'))

        end_time = exam_session['end_time']
        remaining_seconds = max(0, (datetime.fromisoformat(
            end_time) - datetime.now()).total_seconds())

        # Get previous submission if exists
        cursor.execute(
            """
            SELECT * FROM submissions 
            WHERE student_number = ? AND exam_id = ? AND model_id = ?
            ORDER BY submission_time DESC LIMIT 1
            """,
            (session['student_number'],
             session['exam_id'], session['model_id'])
        )
        submission = cursor.fetchone()

        # Initialize with empty answers
        answers = {}
        for question in questions:
            answers[question['id']] = ""

        # If submission exists, get the answers for each question
        if submission:
            # For backwards compatibility, store the old full submission
            prev_code = submission['code_content']

            # Get individual answers for each question
            cursor.execute(
                """
                SELECT question_id, answer_content FROM question_answers
                WHERE submission_id = ?
                """,
                (submission['id'],)
            )
            question_answers = cursor.fetchall()

            for answer in question_answers:
                answers[answer['question_id']] = answer['answer_content']
        else:
            prev_code = ""

        return render_template(
            'exam.html',
            exam=exam,
            model=model,
            questions=questions,
            remaining_seconds=int(remaining_seconds),
            prev_code=prev_code,
            answers=answers
        )

# Auto-save submission API endpoint


@app.route('/api/auto_save', methods=['POST'])
def auto_save():
    if 'student_name' not in session or 'student_number' not in session or 'exam_id' not in session or 'model_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401

    answers = request.json.get('answers', {})  # Get answers for each question
    combined_code = request.json.get(
        'combinedCode', '')  # For backward compatibility
    ip_address = get_real_ip()

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get most recent version number
        cursor.execute(
            """
            SELECT MAX(version) as max_version FROM submissions 
            WHERE student_number = ? AND exam_id = ? AND model_id = ?
            """,
            (session['student_number'],
             session['exam_id'], session['model_id'])
        )
        result = cursor.fetchone()
        new_version = (result[0] or 0) + 1

        # We keep all versions now instead of deleting older ones

        # Insert new submission
        cursor.execute(
            """
            INSERT INTO submissions 
            (student_name, student_number, exam_id, model_id, submission_time, ip_address, code_content, version) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session['student_name'],
                session['student_number'],
                session['exam_id'],
                session['model_id'],
                datetime.now(),
                ip_address,
                combined_code,
                new_version
            )
        )
        submission_id = cursor.lastrowid

        # Save individual answers for each question
        for question_id, answer_content in answers.items():
            cursor.execute(
                """
                INSERT INTO question_answers
                (submission_id, question_id, answer_content)
                VALUES (?, ?, ?)
                """,
                (submission_id, question_id, answer_content)
            )

        conn.commit()
        return jsonify({'success': True, 'version': new_version})

# Final submission endpoint


@app.route('/api/submit', methods=['POST'])
def submit_exam():
    if 'student_name' not in session or 'student_number' not in session or 'exam_id' not in session or 'model_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401

    answers = request.json.get('answers', {})  # Get answers for each question
    combined_code = request.json.get(
        'combinedCode', '')  # For backward compatibility
    ip_address = get_real_ip()

    with sqlite3.connect(DATABASE_PATH) as conn:
        cursor = conn.cursor()

        # Get most recent version number
        cursor.execute(
            """
            SELECT MAX(version) as max_version FROM submissions 
            WHERE student_number = ? AND exam_id = ? AND model_id = ?
            """,
            (session['student_number'],
             session['exam_id'], session['model_id'])
        )
        result = cursor.fetchone()
        new_version = (result[0] or 0) + 1

        # Insert final submission
        cursor.execute(
            """
            INSERT INTO submissions 
            (student_name, student_number, exam_id, model_id, submission_time, ip_address, code_content, version) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session['student_name'],
                session['student_number'],
                session['exam_id'],
                session['model_id'],
                datetime.now(),
                ip_address,
                combined_code,
                new_version
            )
        )
        submission_id = cursor.lastrowid

        # Save individual answers for each question
        for question_id, answer_content in answers.items():
            cursor.execute(
                """
                INSERT INTO question_answers
                (submission_id, question_id, answer_content)
                VALUES (?, ?, ?)
                """,
                (submission_id, question_id, answer_content)
            )

        # Block IP after submission
        cursor.execute(
            """
            UPDATE ip_restrictions 
            SET is_blocked = 1, blocked_time = ?, approved = 0 
            WHERE ip_address = ?
            """,
            (datetime.now(), ip_address)
        )

        # Clear session
        session.pop('student_name', None)
        session.pop('student_number', None)
        session.pop('exam_id', None)
        session.pop('model_id', None)
        session.pop('session_id', None)

        conn.commit()
        return jsonify({'success': True})

# Teacher dashboard route


@app.route('/teacher/dashboard')
def teacher_dashboard():
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM exams ORDER BY id DESC")
        exams = cursor.fetchall()

    return render_template('teacher_dashboard.html', exams=exams)

# Create exam route


@app.route('/teacher/create_exam', methods=['GET', 'POST'])
def create_exam():
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    if request.method == 'POST':
        title = request.form['title']
        duration = int(request.form['duration'])
        model_count = int(request.form.get('model_count', 1))

        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()

            # Insert exam
            cursor.execute(
                "INSERT INTO exams (title, duration, created_by) VALUES (?, ?, ?)",
                (title, duration, session['user_id'])
            )
            exam_id = cursor.lastrowid

            # Create default model (Model A) if no specific models specified
            if model_count == 1:
                cursor.execute(
                    "INSERT INTO exam_models (exam_id, model_name) VALUES (?, ?)",
                    (exam_id, "Model A")
                )
                model_id = cursor.lastrowid

                # Insert questions for default model
                for question in request.form.getlist('question'):
                    if question.strip():
                        cursor.execute(
                            "INSERT INTO questions (exam_id, model_id, question_text) VALUES (?, ?, ?)",
                            (exam_id, model_id, question)
                        )

            conn.commit()

        if model_count > 1:
            # If multiple models specified, redirect to the model creation page
            return redirect(url_for('create_exam_models', exam_id=exam_id, model_count=model_count))

        return redirect(url_for('teacher_dashboard'))

    return render_template('create_exam.html')

# Create exam models route


@app.route('/teacher/create_exam_models/<int:exam_id>', methods=['GET', 'POST'])
def create_exam_models(exam_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    model_count = request.args.get(
        'model_count', type=int) or request.form.get('model_count', type=int)

    if not model_count or model_count < 1:
        model_count = 1

    if request.method == 'POST':
        with sqlite3.connect(DATABASE_PATH) as conn:
            cursor = conn.cursor()

            # Process each model's questions
            for model_num in range(1, model_count + 1):
                # A, B, C, ...
                model_name = request.form.get(
                    f'model_name_{model_num}', f'Model {chr(64 + model_num)}')

                # Insert model
                cursor.execute(
                    "INSERT INTO exam_models (exam_id, model_name) VALUES (?, ?)",
                    (exam_id, model_name)
                )
                model_id = cursor.lastrowid

                # Insert questions for this model
                question_prefix = f'question_{model_num}_'
                for key, value in request.form.items():
                    if key.startswith(question_prefix) and value.strip():
                        question_index = key[len(question_prefix):]
                        cursor.execute(
                            "INSERT INTO questions (exam_id, model_id, question_text) VALUES (?, ?, ?)",
                            (exam_id, model_id, value)
                        )

            conn.commit()

        return redirect(url_for('teacher_dashboard'))

    # For GET request, show form to create models
    with sqlite3.connect(DATABASE_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()

        if not exam:
            return redirect(url_for('teacher_dashboard'))

    return render_template('create_exam_models.html', exam_id=exam_id, model_count=model_count)

# View exam models route


@app.route('/teacher/exam_models/<int:exam_id>')
def view_exam_models(exam_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get exam details
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()

        if not exam:
            return redirect(url_for('teacher_dashboard'))

        # Get all models for this exam
        cursor.execute(
            "SELECT * FROM exam_models WHERE exam_id = ? ORDER BY id", (exam_id,))
        models = cursor.fetchall()

        # Convert models to list of dictionaries for consistent access in template
        models_list = []
        for model in models:
            models_list.append(dict(model))

        # Get questions for each model
        model_questions = {}
        for model in models_list:
            model_id = model['id']
            cursor.execute(
                "SELECT * FROM questions WHERE exam_id = ? AND model_id = ? ORDER BY id",
                (exam_id, model_id)
            )
            questions_rows = cursor.fetchall()

            # Convert questions to list of dictionaries
            questions = []
            for q in questions_rows:
                questions.append(dict(q))

            model_questions[model_id] = questions
            print(
                f"Model {model_id}: {model['model_name']} has {len(questions)} questions")

    return render_template(
        'view_exam_models.html',
        exam=exam,
        models=models_list,
        model_questions=model_questions
    )

# Activate exam route


@app.route('/teacher/activate_exam/<int:exam_id>')
def activate_exam(exam_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        cursor = conn.cursor()

        # Deactivate all exams
        cursor.execute("UPDATE exams SET is_active = 0")

        # Activate selected exam
        cursor.execute(
            "UPDATE exams SET is_active = 1 WHERE id = ?", (exam_id,))

        conn.commit()

    return redirect(url_for('teacher_dashboard'))

# View submissions route


@app.route('/teacher/submissions/<int:exam_id>')
def view_submissions(exam_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get exam details
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()

        # Get all models for this exam
        cursor.execute(
            "SELECT * FROM exam_models WHERE exam_id = ?", (exam_id,))
        exam_models = cursor.fetchall()

        # Convert to dictionary for easy lookup in template
        models = {model['id']: model for model in exam_models}

        # Get all unique students who submitted
        cursor.execute(
            """
            SELECT DISTINCT student_name, student_number, ip_address 
            FROM submissions 
            WHERE exam_id = ?
            """,
            (exam_id,)
        )
        students = cursor.fetchall()

        # Get latest submissions for each student (now showing all instead of just 3)
        student_submissions = {}
        for student in students:
            cursor.execute(
                """
                SELECT * FROM submissions 
                WHERE student_number = ? AND exam_id = ? 
                ORDER BY submission_time DESC
                """,
                (student['student_number'], exam_id)
            )
            submissions = cursor.fetchall()

            # Mark the first one (most recent) as the latest
            if submissions:
                # Convert submissions to list to allow modifications
                submissions_list = []
                for i, sub in enumerate(submissions):
                    # Convert Row to dict
                    sub_dict = dict(sub)
                    # Add is_latest flag
                    # Only the first (most recent) is latest
                    sub_dict['is_latest'] = (i == 0)
                    submissions_list.append(sub_dict)

                student_submissions[student['student_number']
                                    ] = submissions_list
            else:
                student_submissions[student['student_number']] = []

    return render_template(
        'submissions.html',
        exam=exam,
        students=students,
        student_submissions=student_submissions,
        models=models
    )

# Grade submission route


@app.route('/teacher/grade/<int:submission_id>', methods=['GET', 'POST'])
def grade_submission(submission_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get submission details
        cursor.execute("SELECT * FROM submissions WHERE id = ?",
                       (submission_id,))
        submission = cursor.fetchone()

        if not submission:
            return redirect(url_for('teacher_dashboard'))

        # Get the model info for this submission
        cursor.execute("SELECT * FROM exam_models WHERE id = ?",
                       (submission['model_id'],))
        model = cursor.fetchone()

        # Check if this is the latest submission for this student/exam
        cursor.execute(
            """
            SELECT id FROM submissions 
            WHERE student_number = ? AND exam_id = ? 
            ORDER BY submission_time DESC LIMIT 1
            """,
            (submission['student_number'], submission['exam_id'])
        )
        latest_submission = cursor.fetchone()
        is_latest = latest_submission and latest_submission['id'] == submission_id

        # For POST requests (when submitting a grade)
        if request.method == 'POST':
            # Only allow grading if this is the latest submission
            if not is_latest:
                return render_template(
                    'grade.html',
                    submission=submission,
                    model=model,
                    error_message="Only the latest submission can be graded. This is not the latest submission.",
                    is_latest=is_latest,
                    can_grade=False
                )

            mark = float(request.form['mark'])
            comment = request.form['comment']

            # Check if grade already exists
            cursor.execute(
                "SELECT * FROM grades WHERE submission_id = ?", (submission_id,))
            existing_grade = cursor.fetchone()

            if existing_grade:
                cursor.execute(
                    """
                    UPDATE grades 
                    SET mark = ?, comment = ?, graded_by = ?, graded_at = ? 
                    WHERE submission_id = ?
                    """,
                    (mark, comment, session['user_id'],
                     datetime.now(), submission_id)
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO grades 
                    (submission_id, mark, comment, graded_by, graded_at) 
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (submission_id, mark, comment,
                     session['user_id'], datetime.now())
                )

            return redirect(url_for('view_submissions', exam_id=submission['exam_id']))

        # Get existing grade if any
        cursor.execute(
            "SELECT * FROM grades WHERE submission_id = ?", (submission_id,))
        grade = cursor.fetchone()

        # Get exam questions for this specific model
        cursor.execute(
            """
            SELECT * FROM questions 
            WHERE exam_id = ? AND model_id = ?
            ORDER BY id
            """,
            (submission['exam_id'], submission['model_id'])
        )
        questions = cursor.fetchall()

        # Get individual question answers
        cursor.execute(
            """
            SELECT question_id, answer_content FROM question_answers
            WHERE submission_id = ?
            """,
            (submission_id,)
        )
        question_answers_rows = cursor.fetchall()

        # Convert to dictionary for easier access in template
        question_answers = {}
        for row in question_answers_rows:
            question_answers[row['question_id']] = row['answer_content']

    return render_template(
        'grade.html',
        submission=submission,
        model=model,
        grade=grade,
        questions=questions,
        question_answers=question_answers,
        is_latest=is_latest,
        can_grade=is_latest
    )

# IP management route


@app.route('/teacher/ip_management')
def ip_management():
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get all IP restrictions
        cursor.execute("SELECT * FROM ip_restrictions ORDER BY id DESC")
        ip_restrictions = cursor.fetchall()

        # Convert to list of dictionaries for easier manipulation
        ip_data = []
        for ip in ip_restrictions:
            ip_dict = dict(ip)

            # Get the last student who logged in with this IP
            cursor.execute("""
                SELECT student_name, student_number, submission_time 
                FROM submissions 
                WHERE ip_address = ? 
                ORDER BY submission_time DESC 
                LIMIT 1
            """, (ip['ip_address'],))

            last_student = cursor.fetchone()
            if last_student:
                ip_dict['last_student_name'] = last_student['student_name']
                ip_dict['last_student_number'] = last_student['student_number']
                ip_dict['last_login_time'] = last_student['submission_time']
            else:
                # Also check exam_sessions for cases where students logged in but didn't submit
                cursor.execute("""
                    SELECT student_name, student_number, start_time 
                    FROM exam_sessions 
                    WHERE ip_address = ? 
                    ORDER BY start_time DESC 
                    LIMIT 1
                """, (ip['ip_address'],))

                last_session = cursor.fetchone()
                if last_session:
                    ip_dict['last_student_name'] = last_session['student_name']
                    ip_dict['last_student_number'] = last_session['student_number']
                    ip_dict['last_login_time'] = last_session['start_time']
                else:
                    ip_dict['last_student_name'] = None
                    ip_dict['last_student_number'] = None
                    ip_dict['last_login_time'] = None

            ip_data.append(ip_dict)

    return render_template('ip_management.html', ip_restrictions=ip_data)

# Approve IP route


@app.route('/teacher/approve_ip/<int:ip_id>')
def approve_ip(ip_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE ip_restrictions SET is_blocked = 0, approved = 1 WHERE id = ?",
            (ip_id,)
        )
        conn.commit()

    return redirect(url_for('ip_management'))

# Block IP route


@app.route('/teacher/block_ip/<int:ip_id>')
def block_ip(ip_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE ip_restrictions SET is_blocked = 1, blocked_time = ?, approved = 0 WHERE id = ?",
            (datetime.now(), ip_id)
        )
        conn.commit()

    return redirect(url_for('ip_management'))

# View all student grades route


@app.route('/teacher/all_grades')
def all_grades():
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get all exams
        cursor.execute("SELECT * FROM exams ORDER BY id DESC")
        exams = cursor.fetchall()

        # Get all students who have submissions
        cursor.execute(
            """
            SELECT DISTINCT student_name, student_number 
            FROM submissions 
            ORDER BY student_name
            """
        )
        students = cursor.fetchall()

        # Get all exam models
        cursor.execute("SELECT * FROM exam_models")
        exam_models_rows = cursor.fetchall()

        # Convert to dictionary for easy lookup
        exam_models = {}
        for model in exam_models_rows:
            if model['exam_id'] not in exam_models:
                exam_models[model['exam_id']] = {}
            exam_models[model['exam_id']][model['id']] = model['model_name']

        # Get grades only from the latest submissions for each student/exam
        cursor.execute(
            """
            WITH latest_submissions AS (
                SELECT s.id, s.student_number, s.exam_id, s.student_name, s.model_id,
                       ROW_NUMBER() OVER (PARTITION BY s.student_number, s.exam_id 
                                         ORDER BY s.submission_time DESC) AS submission_rank
                FROM submissions s
            )
            SELECT g.*, ls.student_name, ls.student_number, ls.exam_id, ls.model_id, e.title as exam_title
            FROM grades g
            JOIN latest_submissions ls ON g.submission_id = ls.id AND ls.submission_rank = 1
            JOIN exams e ON ls.exam_id = e.id
            ORDER BY e.id DESC, ls.student_name ASC
            """
        )
        all_grades = cursor.fetchall()

        # Organize grades by exam and student for easier display
        grades_by_exam = {}
        for exam in exams:
            grades_by_exam[exam['id']] = {
                'exam_title': exam['title'],
                'students': {}
            }

        for grade in all_grades:
            exam_id = grade['exam_id']
            student_number = grade['student_number']
            model_id = grade['model_id']

            # Get model name if available
            model_name = None
            if exam_id in exam_models and model_id in exam_models[exam_id]:
                model_name = exam_models[exam_id][model_id]

            if exam_id in grades_by_exam:
                if student_number not in grades_by_exam[exam_id]['students']:
                    grades_by_exam[exam_id]['students'][student_number] = {
                        'student_name': grade['student_name'],
                        'student_number': student_number,
                        'grade': grade['mark'],
                        'comment': grade['comment'],
                        'model_name': model_name
                    }

        # Get list of students who haven't been graded for each exam (only check latest submissions)
        # Now include the submission id for direct linking
        ungraded_students = {}
        for exam in exams:
            cursor.execute(
                """
                WITH latest_submissions AS (
                    SELECT s.id, s.student_name, s.student_number, s.model_id,
                           ROW_NUMBER() OVER (PARTITION BY s.student_number, s.exam_id 
                                             ORDER BY s.submission_time DESC) AS submission_rank
                    FROM submissions s
                    WHERE s.exam_id = ?
                )
                SELECT ls.id as submission_id, ls.student_name, ls.student_number, ls.model_id
                FROM latest_submissions ls
                LEFT JOIN grades g ON g.submission_id = ls.id
                WHERE ls.submission_rank = 1 AND g.id IS NULL
                ORDER BY ls.student_name
                """,
                (exam['id'],)
            )
            ungraded_rows = cursor.fetchall()

            # Add model information to ungraded students
            ungraded_with_models = []
            for student in ungraded_rows:
                student_dict = dict(student)

                # Get model name if available
                model_name = None
                model_id = student['model_id']
                if exam['id'] in exam_models and model_id in exam_models[exam['id']]:
                    model_name = exam_models[exam['id']][model_id]

                student_dict['model_name'] = model_name
                ungraded_with_models.append(student_dict)

            ungraded_students[exam['id']] = ungraded_with_models

    return render_template(
        'admin_grades.html',
        exams=exams,
        students=students,
        grades_by_exam=grades_by_exam,
        ungraded_students=ungraded_students,
        exam_models=exam_models
    )

# Export grades to Excel route


@app.route('/teacher/export_grades')
def export_grades():
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get all exams
        cursor.execute("SELECT * FROM exams ORDER BY id DESC")
        exams = cursor.fetchall()

        # Get grades only from the latest submissions for each student/exam
        cursor.execute(
            """
            WITH latest_submissions AS (
                SELECT s.id, s.student_number, s.exam_id, s.student_name,
                       ROW_NUMBER() OVER (PARTITION BY s.student_number, s.exam_id 
                                         ORDER BY s.submission_time DESC) AS submission_rank
                FROM submissions s
            )
            SELECT g.*, ls.student_name, ls.student_number, ls.exam_id, e.title as exam_title
            FROM grades g
            JOIN latest_submissions ls ON g.submission_id = ls.id AND ls.submission_rank = 1
            JOIN exams e ON ls.exam_id = e.id
            ORDER BY e.id DESC, ls.student_name ASC
            """
        )
        all_grades = cursor.fetchall()

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header
    headers = ["Exam Title", "Student Name",
               "Student Number", "Grade", "Comment"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_num, grade in enumerate(all_grades, 2):
        ws.cell(row=row_num, column=1,
                value=grade['exam_title']).border = thin_border
        ws.cell(row=row_num, column=2,
                value=grade['student_name']).border = thin_border
        ws.cell(row=row_num, column=3,
                value=grade['student_number']).border = thin_border
        ws.cell(row=row_num, column=4,
                value=grade['mark']).border = thin_border
        ws.cell(row=row_num, column=5,
                value=grade['comment']).border = thin_border

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Send the file to the user
    return send_file(output, as_attachment=True, download_name="grades.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Export grades to Excel for a specific exam


@app.route('/teacher/export_grades/<int:exam_id>')
def export_grades_excel(exam_id):
    if 'role' not in session or session['role'] != 'teacher':
        return redirect(url_for('login'))

    with sqlite3.connect(DATABASE_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get exam details
        cursor.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cursor.fetchone()

        if not exam:
            return redirect(url_for('teacher_dashboard'))

        # Get all questions for this exam
        cursor.execute(
            "SELECT * FROM questions WHERE exam_id = ? ORDER BY id", (exam_id,))
        questions = cursor.fetchall()

        # Get all models for this exam
        cursor.execute(
            "SELECT * FROM exam_models WHERE exam_id = ?", (exam_id,))
        models_data = cursor.fetchall()

        # Create a model lookup dictionary
        models = {model['id']: model['model_name'] for model in models_data}

        # Get grades only from the latest submissions for each student for this specific exam
        cursor.execute(
            """
            WITH latest_submissions AS (
                SELECT s.id, s.student_number, s.student_name, s.submission_time, s.ip_address, s.model_id,
                       ROW_NUMBER() OVER (PARTITION BY s.student_number 
                                         ORDER BY s.submission_time DESC) AS submission_rank
                FROM submissions s
                WHERE s.exam_id = ?
            )
            SELECT 
                ls.id AS submission_id,
                ls.student_name, 
                ls.student_number, 
                ls.submission_time,
                ls.ip_address,
                ls.model_id,
                g.mark, 
                g.comment, 
                g.graded_at
            FROM latest_submissions ls
            LEFT JOIN grades g ON g.submission_id = ls.id
            WHERE ls.submission_rank = 1
            ORDER BY ls.student_name
            """,
            (exam_id,)
        )
        student_data = cursor.fetchall()

        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        # Excel worksheet names are limited to 31 chars
        ws.title = exam['title'][:31]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        # Helwan University blue
        header_fill = PatternFill("solid", fgColor="183A64")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write exam information
        ws.cell(row=1, column=1, value="Exam:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=exam['title'])
        ws.cell(row=2, column=1, value="Duration (minutes):").font = Font(bold=True)
        ws.cell(row=2, column=2, value=exam['duration'])
        ws.cell(row=3, column=1, value="Export Date:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"))

        # Add some space
        current_row = 5

        # Write headers
        headers = ["Student Name", "Student Number", "Model",
                   "IP Address", "Submission Time", "Grade", "Comment"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set column widths
        column_widths = [20, 15, 15, 15, 20, 10, 30]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[chr(65 + i)].width = width

        # Write student data
        for row_num, student in enumerate(student_data, current_row + 1):
            ws.cell(row=row_num, column=1,
                    value=student['student_name']).border = thin_border
            ws.cell(row=row_num, column=2,
                    value=student['student_number']).border = thin_border

            # Add model information
            model_cell = ws.cell(row=row_num, column=3,
                                 value=models.get(student['model_id'], "Default"))
            model_cell.border = thin_border

            # Apply color to model cell based on model name
            model_name = models.get(student['model_id'], "").upper()
            if "A" in model_name:
                model_cell.fill = PatternFill(
                    "solid", fgColor="DCE6F1")  # Light blue
            elif "B" in model_name:
                model_cell.fill = PatternFill(
                    "solid", fgColor="E6DCF1")  # Light purple
            elif "C" in model_name:
                model_cell.fill = PatternFill(
                    "solid", fgColor="DCF1E6")  # Light green
            elif "D" in model_name:
                model_cell.fill = PatternFill(
                    "solid", fgColor="F1DCE6")  # Light pink

            ws.cell(row=row_num, column=4,
                    value=student['ip_address']).border = thin_border
            ws.cell(row=row_num, column=5,
                    value=student['submission_time']).border = thin_border

            # Grade might be NULL if not graded yet
            grade_cell = ws.cell(
                row=row_num, column=6, value=student['mark'] if student['mark'] is not None else "Not graded")
            grade_cell.border = thin_border

            # If not graded, add red fill
            if student['mark'] is None:
                grade_cell.fill = PatternFill("solid", fgColor="FFCCCC")

            comment_cell = ws.cell(
                row=row_num, column=7, value=student['comment'] if student['comment'] is not None else "")
            comment_cell.border = thin_border

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate a meaningful filename with the exam title and date
        safe_title = ''.join(
            c for c in exam['title'] if c.isalnum() or c in ' _-').strip()
        safe_title = safe_title.replace(' ', '_')
        filename = f"{safe_title}_grades_{datetime.now().strftime('%Y%m%d')}.xlsx"

        # Send the file to the user
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Logout route


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
